from __future__ import annotations

import io as py_io
from datetime import date, datetime
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from klassenbildung.core.constants import (
    BASIS_SHEET_NAME,
    COLUMN_BY_INDEX,
    HEADER_SCAN_ROWS,
    MUSIC_FLAG_COLUMNS,
)
from klassenbildung.core.models import ImportResult, Student, ValidationMessage
from klassenbildung.core.normalization import (
    normalize_class_id,
    normalize_eligibility,
    normalize_gender,
    normalize_language,
    normalize_music_profile,
    normalize_string,
)


def import_excel(source: bytes | BinaryIO, filename: str | None = None) -> ImportResult:
    workbook_bytes = _read_bytes(source)
    try:
        workbook = load_workbook(py_io.BytesIO(workbook_bytes), data_only=False)
    except Exception as exc:  # pragma: no cover - exact openpyxl errors vary
        return ImportResult(
            students=[],
            messages=[ValidationMessage("FEHLER", f"Excel-Datei konnte nicht gelesen werden: {exc}")],
            sheet_names=[],
            detected_classes=[],
            workbook_bytes=workbook_bytes,
            source_filename=filename,
        )

    sheet_names = list(workbook.sheetnames)
    if BASIS_SHEET_NAME not in workbook.sheetnames:
        return ImportResult(
            students=[],
            messages=[ValidationMessage("FEHLER", 'Blatt "Basis" fehlt.')],
            sheet_names=sheet_names,
            detected_classes=[],
            workbook_bytes=workbook_bytes,
            source_filename=filename,
        )

    sheet = workbook[BASIS_SHEET_NAME]
    header_row = find_header_row(sheet)
    if header_row is None:
        return ImportResult(
            students=[],
            messages=[ValidationMessage("FEHLER", "Header-Zeile wurde nicht erkannt.")],
            sheet_names=sheet_names,
            detected_classes=[],
            workbook_bytes=workbook_bytes,
            source_filename=filename,
        )

    students = parse_students(sheet, header_row)
    detected_classes = sorted(
        {student.original_class for student in students if student.original_class},
        key=_class_sort_key,
    )
    messages = [
        ValidationMessage("INFO", f"{len(students)} Schüler erkannt."),
        ValidationMessage("INFO", f"{len(detected_classes)} Klassen erkannt."),
        ValidationMessage(
            "INFO",
            f"{sum(1 for student in students if student.comment)} Bemerkungen gefunden.",
        ),
    ]

    return ImportResult(
        students=students,
        messages=messages,
        sheet_names=sheet_names,
        detected_classes=detected_classes,
        workbook_bytes=workbook_bytes,
        source_filename=filename,
    )


def _read_bytes(source: bytes | BinaryIO) -> bytes:
    if isinstance(source, bytes):
        return source
    position = None
    if hasattr(source, "tell"):
        position = source.tell()
    data = source.read()
    if position is not None and hasattr(source, "seek"):
        source.seek(position)
    return data


def find_header_row(sheet: Worksheet) -> int | None:
    for row_index in range(1, min(sheet.max_row, HEADER_SCAN_ROWS) + 1):
        values = [
            normalize_string(sheet.cell(row=row_index, column=column).value) or ""
            for column in range(1, 21)
        ]
        lowered = {value.lower() for value in values}
        if "nr" in lowered and "abgebende schule" in lowered:
            return row_index
    return 1 if sheet.max_row else None


def parse_students(sheet: Worksheet, header_row: int) -> list[Student]:
    students: list[Student] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        raw = {
            field_name: sheet.cell(row=row_number, column=column_index).value
            for column_index, field_name in COLUMN_BY_INDEX.items()
        }
        if not is_student_row(raw):
            continue
        student = build_student(row_number, raw)
        students.append(student)
    return students


def is_student_row(raw: dict[str, Any]) -> bool:
    school = normalize_string(raw.get("school"))
    if not school:
        return False
    has_student_marker = any(
        normalize_string(raw.get(field_name))
        for field_name in ("nr", "gender", "second_language", "music_profile")
    )
    has_music_flag = any(normalize_string(raw.get(field_name)) for field_name in MUSIC_FLAG_COLUMNS)
    return has_student_marker or has_music_flag


def build_student(row_number: int, raw: dict[str, Any]) -> Student:
    eligibility = normalize_eligibility(raw.get("eligibility"))
    music_profile = normalize_music_profile(raw.get("music_profile")) or derive_music_profile(raw)
    nr = normalize_string(raw.get("nr"))
    birthdate = _normalize_birthdate(raw.get("birthdate"))
    return Student(
        internal_id=f"row-{row_number}",
        row_number=row_number,
        original_class=normalize_class_id(raw.get("zielklasse")),
        nr=nr,
        school=normalize_string(raw.get("school")),
        last_name=normalize_string(raw.get("last_name")),
        first_name=normalize_string(raw.get("first_name")),
        eligibility=eligibility,
        gender=normalize_gender(raw.get("gender")),
        birthdate=birthdate,
        nationality=normalize_string(raw.get("nationality")),
        religion=normalize_string(raw.get("religion")),
        second_language=normalize_language(raw.get("second_language")),
        music_profile=music_profile,
        primary_class=normalize_string(raw.get("primary_class")),
        friend1=normalize_string(raw.get("friend1")),
        friend2=normalize_string(raw.get("friend2")),
        comment=normalize_string(raw.get("comment")),
        is_support=eligibility == "R",
    )


def derive_music_profile(raw: dict[str, Any]) -> str | None:
    for field_name, profile in MUSIC_FLAG_COLUMNS.items():
        value = normalize_string(raw.get(field_name))
        if value:
            return profile
    return None


def _normalize_birthdate(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _class_sort_key(class_id: str) -> tuple[int, str]:
    number = "".join(character for character in class_id if character.isdigit())
    letter = "".join(character for character in class_id if character.isalpha())
    return (int(number) if number else 0, letter)

