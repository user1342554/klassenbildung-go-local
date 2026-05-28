from __future__ import annotations

import io as py_io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from klassenbildung.core.constants import BASIS_SHEET_NAME, EXPORT_COLUMN_COUNT
from klassenbildung.core.models import ClassConfig, ScoreReport, Student, ValidationMessage


def export_excel(
    source_workbook_bytes: bytes | None,
    students: list[Student],
    assignments: dict[str, str],
    class_configs: list[ClassConfig],
    score_report: ScoreReport | None = None,
    validation_messages: list[ValidationMessage] | None = None,
) -> bytes:
    workbook = _load_or_create_workbook(source_workbook_bytes)
    basis = workbook[BASIS_SHEET_NAME]

    for student in students:
        class_id = assignments.get(student.internal_id)
        if class_id:
            basis.cell(row=student.row_number, column=1).value = class_id

    _replace_class_sheets(workbook, basis, students, assignments, class_configs)
    _write_score_sheet(workbook, score_report)
    _write_warning_sheet(workbook, validation_messages or [])

    output = py_io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_or_create_workbook(source_workbook_bytes: bytes | None) -> Workbook:
    if source_workbook_bytes:
        return load_workbook(py_io.BytesIO(source_workbook_bytes))
    workbook = Workbook()
    workbook.active.title = BASIS_SHEET_NAME
    return workbook


def _replace_class_sheets(
    workbook: Workbook,
    basis,
    students: list[Student],
    assignments: dict[str, str],
    class_configs: list[ClassConfig],
) -> None:
    class_ids = {config.class_id for config in class_configs}
    for title in list(workbook.sheetnames):
        if title in class_ids:
            del workbook[title]

    header_values = [basis.cell(row=1, column=column).value for column in range(1, EXPORT_COLUMN_COUNT + 1)]
    students_by_class = {config.class_id: [] for config in class_configs}
    for student in students:
        class_id = assignments.get(student.internal_id)
        if class_id in students_by_class:
            students_by_class[class_id].append(student)

    for config in class_configs:
        sheet = workbook.create_sheet(config.class_id)
        sheet.append(header_values)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for student in sorted(students_by_class[config.class_id], key=lambda item: (item.sort_name, item.row_number)):
            values = [
                basis.cell(row=student.row_number, column=column).value
                for column in range(1, EXPORT_COLUMN_COUNT + 1)
            ]
            values[0] = config.class_id
            sheet.append(values)
        for column in range(1, EXPORT_COLUMN_COUNT + 1):
            letter = sheet.cell(row=1, column=column).column_letter
            sheet.column_dimensions[letter].width = basis.column_dimensions[letter].width or 12


def _write_score_sheet(workbook: Workbook, score_report: ScoreReport | None) -> None:
    if "Auswertung" in workbook.sheetnames:
        del workbook["Auswertung"]
    sheet = workbook.create_sheet("Auswertung")
    sheet.append(["Kennzahl", "Wert"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    if not score_report:
        sheet.append(["Status", "Keine Auswertung vorhanden"])
        return
    sheet.append(["Gesamtscore", score_report.total_score])
    sheet.append(["Harte Regelverletzungen", len(score_report.hard_violations)])
    sheet.append(["Freund 1 erfüllt", f"{score_report.friend1_fulfilled}/{score_report.friend1_total}"])
    sheet.append(["Freund 2 erfüllt", f"{score_report.friend2_fulfilled}/{score_report.friend2_total}"])
    sheet.append([])
    sheet.append(["Klasse", "Anzahl", "m", "w", "F", "L", "Reg", "B", "S", "G", "R"])
    for report in score_report.class_reports:
        sheet.append(
            [
                report.class_id,
                report.size,
                report.gender_counts.get("m", 0),
                report.gender_counts.get("w", 0),
                report.language_counts.get("F", 0),
                report.language_counts.get("L", 0),
                report.music_counts.get("Reg", 0),
                report.music_counts.get("B", 0),
                report.music_counts.get("S", 0),
                report.music_counts.get("G", 0),
                report.support_count,
            ]
        )


def _write_warning_sheet(workbook: Workbook, messages: list[ValidationMessage]) -> None:
    if "Warnungen" in workbook.sheetnames:
        del workbook["Warnungen"]
    sheet = workbook.create_sheet("Warnungen")
    sheet.append(["Typ", "Zeile", "Spalte", "Wert", "Meldung"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for message in messages:
        sheet.append(
            [
                message.severity,
                message.row_number,
                message.column,
                message.value,
                message.message,
            ]
        )

