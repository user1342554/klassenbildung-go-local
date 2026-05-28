from __future__ import annotations

import io

from openpyxl import load_workbook

from klassenbildung.core.models import ClassConfig
from klassenbildung.core.settings import load_settings
from klassenbildung.excel_io.excel_export import export_excel
from klassenbildung.excel_io.excel_import import import_excel
from klassenbildung.optimization.scoring import score_solution


def test_export_updates_basis_and_creates_class_sheets(sample_workbook_bytes: bytes) -> None:
    result = import_excel(sample_workbook_bytes)
    class_configs = [
        ClassConfig("5a", "5a", 0, 3, [], []),
        ClassConfig("5b", "5b", 0, 3, [], []),
    ]
    assignments = {
        result.students[0].internal_id: "5a",
        result.students[1].internal_id: "5b",
        result.students[2].internal_id: "5b",
    }
    score = score_solution(result.students, assignments, load_settings(), class_configs)

    exported = export_excel(sample_workbook_bytes, result.students, assignments, class_configs, score, [])
    workbook = load_workbook(io.BytesIO(exported))

    assert "Basis" in workbook.sheetnames
    assert "5a" in workbook.sheetnames
    assert "5b" in workbook.sheetnames
    assert "Auswertung" in workbook.sheetnames
    assert "Warnungen" in workbook.sheetnames
    assert workbook["Basis"]["A3"].value == "5b"
    assert workbook["5b"].max_row == 3

